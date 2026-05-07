import re
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

_RED    = PatternFill(start_color="FF746C", end_color="FF746C", fill_type="solid")
_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_GREEN  = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
_ORANGE = PatternFill(start_color="FFD099", end_color="FFD099", fill_type="solid")

_NAME_SIMILARITY_THRESHOLD = 0.9

_LEGEND = """\
LEGENDA
-------
ROOD   = Ontbrekend veld of geen/ongeldige RecordID
GEEL   = Mogelijke duplicaat op basis van RecordID
ORANJE = Onverwachte waarde in RecordID
GROEN  = Mogelijke duplicaat op basis van naam

Opmerking: als een rij zowel op naamovereenkomst als op een RecordID-regel valt,
           heeft de RecordID-kleur voorrang in het Excel-bestand.
           Ontbrekende velden (ROOD) hebben altijd voorrang op alle andere kleuren.

"""

_COLOR_NL = {"RED": "ROOD", "YELLOW": "GEEL", "GREEN": "GROEN", "ORANGE": "ORANJE"}


class ExcelFormatter:
    def format(self, path, processing_log=None):
        wb = load_workbook(path)
        ws = wb.active
        log = defaultdict(lambda: {'color': None, 'reasons': []})
        self._apply_name_similarity_rules(ws, log)
        self._apply_record_id_rules(ws, log)
        self._apply_missing_value_rules(ws, log)  # runs last — RED always wins
        wb.save(path)
        self._write_explain_log(path, log, processing_log or [])

    # --- row rules ---

    def _apply_name_similarity_rules(self, ws, log):
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "Name" not in header:
            return

        col = header.index("Name")
        rows = list(ws.iter_rows(min_row=2))
        names = [str(row[col].value or "") for row in rows]

        similar = defaultdict(list)
        for i, name_a in enumerate(names):
            if not name_a:
                continue
            for j, name_b in enumerate(names):
                if i >= j or not name_b:
                    continue
                ratio = SequenceMatcher(None, name_a, name_b).ratio()
                if ratio >= _NAME_SIMILARITY_THRESHOLD:
                    similar[i].append((j, ratio))
                    similar[j].append((i, ratio))

        for i, matches in similar.items():
            for cell in rows[i]:
                cell.fill = _GREEN
            row_num = i + 2
            match_strs = [f"rij {j + 2} \"{names[j]}\" ({ratio:.0%})" for j, ratio in matches]
            log[row_num]['color'] = 'GREEN'
            log[row_num]['reasons'].append(
                f"Naam \"{names[i]}\" lijkt op: {', '.join(match_strs)}"
            )

    def _apply_record_id_rules(self, ws, log):
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "RecordID" not in header:
            return

        col = header.index("RecordID")
        values = [str(row[col].value or "") for row in ws.iter_rows(min_row=2)]
        counts = Counter(v for v in values if v)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            value = str(row[col].value or "")
            numbers = re.findall(r'\d+', value)

            if not value:
                fill, color = _RED, 'RED'
                reason = "RecordID is leeg"
            elif len(numbers) == 2:
                fill, color = _RED, 'RED'
                reason = f"RecordID \"{value}\" bevat twee nummergroepen (mogelijke samengevoegde invoer)"
            elif counts[value] > 1:
                fill, color = _YELLOW, 'YELLOW'
                reason = f"RecordID \"{value}\" komt {counts[value]} keer voor op deze pagina"
            elif len(numbers) == 0:
                fill, color = _ORANGE, 'ORANGE'
                reason = f"RecordID \"{value}\" bevat geen cijfers"
            else:
                continue

            for cell in row:
                cell.fill = fill
            row_num = row_idx + 2
            log[row_num]['color'] = color  # overwrites GREEN when a RecordID rule fires
            log[row_num]['reasons'].append(reason)

    def _apply_missing_value_rules(self, ws, log):
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        _SKIP = {"Full Text", "RecordID", "ExtraInformation", "Delimiter"}  # RecordID has its own rule; others are optional
        check_cols = [i for i, h in enumerate(header) if h and h not in _SKIP]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            missing = [header[i] for i in check_cols if not str(row[i].value or "").strip()]
            if not missing:
                continue
            for cell in row:
                cell.fill = _RED
            row_num = row_idx + 2
            log[row_num]['color'] = 'RED'
            for col_name in missing:
                log[row_num]['reasons'].append(f"Kolom '{col_name}' is leeg")

    # --- explain log ---

    def _write_explain_log(self, excel_path, log, processing_log):
        log_path = Path(excel_path).with_name(Path(excel_path).stem + "_explain.txt")
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(_LEGEND)
            if processing_log:
                f.write("VERWERKINGSLOG\n")
                f.write("--------------\n")
                for msg in processing_log:
                    f.write(f"{msg}\n")
                f.write("\n")
            f.write("GEKLEURDE RIJEN\n")
            f.write("---------------\n")
            for row_num in sorted(log):
                entry = log[row_num]
                color_nl = _COLOR_NL.get(entry['color'], entry['color'])
                for reason in entry['reasons']:
                    f.write(f"Rij {row_num:4d}  [{color_nl:6}]  {reason}\n")
